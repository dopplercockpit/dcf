"""
Excel workbook builder for DCF outputs (openpyxl).
"""

from datetime import datetime
from io import BytesIO
import json

from openpyxl import Workbook


def _safe_get(mapping, key, default=None):
    if isinstance(mapping, dict):
        return mapping.get(key, default)
    return default


def _append_rows(sheet, rows):
    for row in rows:
        sheet.append(list(row))


def build_workbook_from_results(ticker, results):
    company = _safe_get(results, "company_data", {})
    assumptions = _safe_get(results, "assumptions", {})
    wacc_results = _safe_get(results, "wacc_results", {})
    esg = _safe_get(results, "esg", {})
    stress = _safe_get(results, "stress_test", {})

    projected_fcf = _safe_get(results, "projected_fcf", []) or []
    pv_fcf = _safe_get(results, "pv_fcf", []) or []
    terminal_value = _safe_get(results, "terminal_value", 0)
    pv_terminal_value = _safe_get(results, "pv_terminal_value", 0)

    wb = Workbook()
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"

    inputs_rows = [
        ("Ticker", ticker),
        ("Company Name", _safe_get(company, "company_name")),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Tax Rate", _safe_get(assumptions, "tax_rate")),
        ("Risk-Free Rate", _safe_get(assumptions, "risk_free_rate")),
        ("Market Risk Premium", _safe_get(assumptions, "market_risk_premium")),
        ("Beta", _safe_get(assumptions, "beta")),
        ("Cost of Debt", _safe_get(assumptions, "cost_of_debt")),
        ("Perpetual Growth Rate", _safe_get(assumptions, "perpetual_growth_rate")),
        ("Revenue Growth Rates", json.dumps(_safe_get(assumptions, "revenue_growth_rates", []))),
        ("ESG Adjustment Enabled", _safe_get(assumptions, "esg_adjustment_enabled")),
        ("ESG Strength (bps)", _safe_get(assumptions, "esg_strength_bps")),
        ("ESG Good Threshold", _safe_get(assumptions, "esg_threshold_good")),
        ("ESG Bad Threshold", _safe_get(assumptions, "esg_threshold_bad")),
        ("Stress Enabled", _safe_get(assumptions, "stress_enabled")),
        ("Supply Chain Shock", _safe_get(assumptions, "stress_supply_chain")),
        ("Carbon Tax", _safe_get(assumptions, "stress_carbon_tax")),
        ("Carbon Intensity", _safe_get(assumptions, "carbon_intensity")),
        ("Carbon Tax Rate", _safe_get(assumptions, "carbon_tax_rate")),
        ("ESG Total Score", _safe_get(esg, "total_esg")),
        ("ESG Environment Score", _safe_get(esg, "environment_score")),
        ("ESG Social Score", _safe_get(esg, "social_score")),
        ("ESG Governance Score", _safe_get(esg, "governance_score")),
        ("ESG Controversy Level", _safe_get(esg, "controversy_level")),
    ]
    _append_rows(ws_inputs, inputs_rows)

    ws_forecast = wb.create_sheet("Forecast")
    forecast_header = [
        "Year",
        "FCF Base",
        "FCF Stressed",
        "Carbon Costs",
        "Discount Factor",
        "PV of FCF",
    ]
    ws_forecast.append(forecast_header)

    wacc = _safe_get(wacc_results, "wacc", 0) or 0
    stressed_fcf = _safe_get(stress, "stressed_projected_fcf", []) or []
    carbon_costs = _safe_get(stress, "carbon_costs", []) or []

    for idx, base_fcf in enumerate(projected_fcf):
        year = idx + 1
        discount = 1 / ((1 + wacc) ** year) if wacc is not None else None
        row = [
            year,
            base_fcf,
            stressed_fcf[idx] if idx < len(stressed_fcf) else None,
            carbon_costs[idx] if idx < len(carbon_costs) else None,
            discount,
            pv_fcf[idx] if idx < len(pv_fcf) else None,
        ]
        ws_forecast.append(row)

    ws_forecast.append([])
    ws_forecast.append(["Terminal Value", terminal_value])
    ws_forecast.append(["PV Terminal Value", pv_terminal_value])

    ws_summary = wb.create_sheet("Summary")
    summary_rows = [
        ("WACC", _safe_get(wacc_results, "wacc")),
        ("Cost of Equity (Ke)", _safe_get(wacc_results, "cost_of_equity")),
        ("Ke Before ESG", _safe_get(wacc_results, "ke_before_esg")),
        ("Ke After ESG", _safe_get(wacc_results, "ke_after_esg")),
        ("After-Tax Cost of Debt", _safe_get(wacc_results, "after_tax_cost_debt")),
        ("Equity Weight", _safe_get(wacc_results, "equity_weight")),
        ("Debt Weight", _safe_get(wacc_results, "debt_weight")),
        ("Intrinsic Value per Share (Base)", _safe_get(results, "intrinsic_value_per_share")),
        ("Intrinsic Value per Share (Stressed)", _safe_get(stress, "stressed_intrinsic_value_per_share")),
        ("Current Price", _safe_get(results, "current_market_value")),
        ("Upside %", _safe_get(results, "upside_pct")),
        ("Data Quality", _safe_get(_safe_get(results, "data_quality", {}), "quality")),
    ]
    _append_rows(ws_summary, summary_rows)

    return wb


def build_workbook_bytes(ticker, results):
    workbook = build_workbook_from_results(ticker, results)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
